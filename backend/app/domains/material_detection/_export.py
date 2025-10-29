"""
Material Detection Export Utilities

CSV and Excel export functionality for detection results.
"""

import io
import csv
from typing import List
from datetime import datetime
from .models import DetectedMaterial


def export_materials_to_csv(materials: List[DetectedMaterial]) -> io.StringIO:
    """
    Export detected materials to CSV format.

    Args:
        materials: List of DetectedMaterial objects

    Returns:
        StringIO buffer containing CSV data
    """
    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    headers = [
        'Material ID',
        'Job ID',
        'Image ID',
        'Category',
        'Type',
        'Grade',
        'Finish',
        'Confidence Score',
        'Provider',
        'Quantity Estimate',
        'Unit Type',
        'Unit Price',
        'Total Estimate',
        'Needs Review',
        'Reviewed By',
        'Review Notes',
        'Created At'
    ]
    writer.writerow(headers)

    # Write data
    for material in materials:
        row = [
            str(material.id),
            str(material.job_id),
            str(material.image_id),
            material.material_category or '',
            material.material_type or '',
            material.material_grade or '',
            material.material_finish or '',
            f"{float(material.confidence_score):.4f}",
            material.provider_used or '',
            str(material.quantity_estimate) if material.quantity_estimate else '',
            material.unit_type or '',
            str(material.unit_price) if material.unit_price else '',
            str(material.total_estimate) if material.total_estimate else '',
            'Yes' if material.needs_review else 'No',
            str(material.reviewed_by_id) if material.reviewed_by_id else '',
            material.review_notes or '',
            material.created_at.isoformat() if material.created_at else ''
        ]
        writer.writerow(row)

    output.seek(0)
    return output


def export_materials_to_excel(materials: List[DetectedMaterial]) -> io.BytesIO:
    """
    Export detected materials to Excel format.

    Args:
        materials: List of DetectedMaterial objects

    Returns:
        BytesIO buffer containing Excel data
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl package required for Excel export")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Detected Materials"

    # Define headers
    headers = [
        'Material ID',
        'Job ID',
        'Image ID',
        'Category',
        'Type',
        'Grade',
        'Finish',
        'Confidence Score',
        'Provider',
        'Quantity Estimate',
        'Unit Type',
        'Unit Price',
        'Total Estimate',
        'Needs Review',
        'Reviewed By',
        'Review Notes',
        'Created At'
    ]

    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Write data
    for row_num, material in enumerate(materials, 2):
        ws.cell(row=row_num, column=1, value=str(material.id))
        ws.cell(row=row_num, column=2, value=str(material.job_id))
        ws.cell(row=row_num, column=3, value=str(material.image_id))
        ws.cell(row=row_num, column=4, value=material.material_category or '')
        ws.cell(row=row_num, column=5, value=material.material_type or '')
        ws.cell(row=row_num, column=6, value=material.material_grade or '')
        ws.cell(row=row_num, column=7, value=material.material_finish or '')
        ws.cell(row=row_num, column=8, value=float(material.confidence_score))
        ws.cell(row=row_num, column=9, value=material.provider_used or '')
        ws.cell(row=row_num, column=10, value=float(material.quantity_estimate) if material.quantity_estimate else None)
        ws.cell(row=row_num, column=11, value=material.unit_type or '')
        ws.cell(row=row_num, column=12, value=float(material.unit_price) if material.unit_price else None)
        ws.cell(row=row_num, column=13, value=float(material.total_estimate) if material.total_estimate else None)
        ws.cell(row=row_num, column=14, value='Yes' if material.needs_review else 'No')
        ws.cell(row=row_num, column=15, value=str(material.reviewed_by_id) if material.reviewed_by_id else '')
        ws.cell(row=row_num, column=16, value=material.review_notes or '')
        ws.cell(row=row_num, column=17, value=material.created_at if material.created_at else '')

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output
