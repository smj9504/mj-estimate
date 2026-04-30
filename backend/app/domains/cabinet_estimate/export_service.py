"""
Cabinet Estimate export service - PDF and Excel generation.
"""

import base64
import io
import logging
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

logger = logging.getLogger(__name__)

# ── Line item category → section mapping ──
SECTION_ORDER = [
    ("Cabinet Supply", ["supply", "premium"]),
    ("Demolition & Removal", ["demo"]),
    ("Installation", ["install"]),
    ("Plumbing & Fixtures", ["plumbing"]),
    ("Countertop", ["countertop"]),
    ("Finishing", ["finishing"]),
    ("General Conditions", ["misc"]),
]


def _group_line_items(
    line_items: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    """Group line items by location then category."""
    sections = []

    has_island = any(
        li.get("location") == "island"
        for li in line_items
    )

    if has_island:
        # Perimeter → Island → Shared
        for loc, loc_label in [
            ("perimeter", "Perimeter"),
            ("island", "Island"),
        ]:
            loc_items = [
                li for li in line_items
                if li.get("location", "perimeter")
                == loc
            ]
            if not loc_items:
                continue
            for title, cats in SECTION_ORDER:
                items = [
                    li for li in loc_items
                    if li.get("category") in cats
                ]
                if items:
                    total = sum(
                        li.get("total", 0)
                        for li in items
                    )
                    sections.append({
                        "title": (
                            f"{loc_label} — {title}"
                        ),
                        "items": items,
                        "total": total,
                    })

        shared = [
            li for li in line_items
            if li.get("location") == "shared"
        ]
        if shared:
            for title, cats in SECTION_ORDER:
                items = [
                    li for li in shared
                    if li.get("category") in cats
                ]
                if items:
                    total = sum(
                        li.get("total", 0)
                        for li in items
                    )
                    sections.append({
                        "title": title,
                        "items": items,
                        "total": total,
                    })
    else:
        for title, cats in SECTION_ORDER:
            items = [
                li for li in line_items
                if li.get("category") in cats
            ]
            if items:
                total = sum(
                    li.get("total", 0)
                    for li in items
                )
                sections.append({
                    "title": title,
                    "items": items,
                    "total": total,
                })
    return sections


class CabinetExportService:
    """Generates PDF estimates and Excel purchase orders."""

    # ── PDF Export ──

    def generate_pdf(
        self,
        estimate: Dict[str, Any],
        show_signature: bool = True,
    ) -> io.BytesIO:
        """Generate a professional PDF estimate."""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter
        from reportlab.lib.styles import (
            ParagraphStyle,
            getSampleStyleSheet,
        )
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            HRFlowable,
            Image,
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        buffer = io.BytesIO()
        page_w, page_h = letter
        margin = 0.6 * inch

        doc = SimpleDocTemplate(
            buffer,
            pagesize=letter,
            rightMargin=margin,
            leftMargin=margin,
            topMargin=margin,
            bottomMargin=0.8 * inch,
        )

        usable_w = page_w - 2 * margin

        # Page header/footer data (used by callbacks)
        est_id_short = str(
            estimate.get("id", "")
        )[:8].upper()
        prop_addr = estimate.get(
            "property_address", ""
        ) or ""

        def _on_later_pages(canvas, doc_ref):
            """Header + footer on pages 2+."""
            canvas.saveState()
            # Header: property address (left) + quote # (right)
            canvas.setFont("Helvetica", 7.5)
            canvas.setFillColor(colors.HexColor("#4a5568"))
            canvas.drawString(
                margin, page_h - margin + 8,
                prop_addr,
            )
            canvas.drawRightString(
                page_w - margin, page_h - margin + 8,
                f"Quote # {est_id_short}",
            )
            # thin line below header
            canvas.setStrokeColor(
                colors.HexColor("#e2e8f0")
            )
            canvas.setLineWidth(0.5)
            canvas.line(
                margin, page_h - margin + 3,
                page_w - margin, page_h - margin + 3,
            )
            # Footer: page number
            canvas.setFont("Helvetica", 7.5)
            canvas.drawCentredString(
                page_w / 2, 0.45 * inch,
                f"Page {doc_ref.page}",
            )
            canvas.restoreState()

        # ── Colors ──
        brand_dark = colors.HexColor("#1a2332")
        brand_accent = colors.HexColor("#2c5282")
        brand_light_bg = colors.HexColor("#f7fafc")
        text_grey = colors.HexColor("#4a5568")
        border_grey = colors.HexColor("#e2e8f0")
        section_bg = colors.HexColor("#edf2f7")

        # ── Styles ──
        styles = getSampleStyleSheet()

        company_info_style = ParagraphStyle(
            "CompanyInfo",
            fontSize=8,
            fontName="Helvetica",
            textColor=text_grey,
            leading=12,
        )
        doc_title_style = ParagraphStyle(
            "DocTitle",
            fontSize=22,
            fontName="Helvetica-Bold",
            textColor=brand_dark,
            spaceAfter=2,
        )
        section_title_style = ParagraphStyle(
            "SectionTitle",
            fontSize=10,
            fontName="Helvetica-Bold",
            textColor=brand_dark,
            spaceBefore=14,
            spaceAfter=4,
        )
        normal = ParagraphStyle(
            "NormalCustom",
            fontSize=9,
            fontName="Helvetica",
            textColor=brand_dark,
        )
        small_grey = ParagraphStyle(
            "SmallGrey",
            fontSize=7.5,
            fontName="Helvetica",
            textColor=text_grey,
            leading=10,
        )
        terms_style = ParagraphStyle(
            "Terms",
            fontSize=8.5,
            fontName="Helvetica",
            textColor=text_grey,
            leading=12,
            spaceBefore=2,
        )

        elements = []

        # ══════════════════════════════════════
        # HEADER: Company info + Estimate title
        # ══════════════════════════════════════

        company = estimate.get("company_info") or {}

        # Left: logo + company info
        left_elements = []

        # Company text (single Paragraph for table cell)
        company_parts = []
        if company.get("name"):
            company_parts.append(
                f"<b>{company['name']}</b>"
            )
        if company.get("address"):
            addr = company["address"]
            city_state = ", ".join(
                filter(None, [
                    company.get("city"),
                    company.get("state"),
                    company.get("zipcode"),
                ])
            )
            if city_state:
                addr += f", {city_state}"
            company_parts.append(addr)
        contact = []
        if company.get("phone"):
            contact.append(company["phone"])
        if company.get("email"):
            contact.append(company["email"])
        if contact:
            company_parts.append(
                " | ".join(contact),
            )
        if company.get("license_number"):
            company_parts.append(
                f"License #: "
                f"{company['license_number']}"
            )

        if company_parts:
            company_cell = Paragraph(
                "<br/>".join(company_parts),
                company_info_style,
            )
        else:
            company_cell = Spacer(1, 1)

        # Right: QUOTE title + number + date
        est_date = datetime.now()
        valid_until = est_date + timedelta(days=30)

        right_text = (
            f"<font size=20><b>QUOTE</b></font>"
            f"<br/><br/>"
            f"<font size=8>"
            f"<b>Quote #:</b> {est_id_short}<br/>"
            f"<b>Date:</b> "
            f"{est_date.strftime('%B %d, %Y')}<br/>"
            f"<b>Valid Until:</b> "
            f"{valid_until.strftime('%B %d, %Y')}"
            f"</font>"
        )
        right_cell = Paragraph(
            right_text,
            ParagraphStyle(
                "RightHeader",
                fontSize=8,
                fontName="Helvetica",
                textColor=brand_dark,
                alignment=2,  # RIGHT
                leading=14,
            ),
        )

        # Build header table
        header_data = [[company_cell, right_cell]]
        header_table = Table(
            header_data,
            colWidths=[usable_w * 0.6, usable_w * 0.4],
        )
        header_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (1, 0), (1, 0), "RIGHT"),
        ]))
        elements.append(header_table)

        elements.append(Spacer(1, 14))

        # ══════════════════════════════════════
        # PROJECT INFO: Bill To + Project
        # ══════════════════════════════════════

        bill_to_parts = []
        if estimate.get("client_name"):
            bill_to_parts.append(
                f"<b>{estimate['client_name']}</b>"
            )
        if estimate.get("property_address"):
            bill_to_parts.append(
                estimate["property_address"]
            )
        if estimate.get("claim_number"):
            bill_to_parts.append(
                f"Claim #: {estimate['claim_number']}"
            )

        project_parts = []
        if estimate.get("layout_type"):
            project_parts.append(
                f"Layout: {estimate['layout_type']}"
            )
        if estimate.get("kitchen_size_sqft"):
            project_parts.append(
                f"Kitchen Size: "
                f"{estimate['kitchen_size_sqft']} sqft"
            )
        tier = estimate.get("tier", "")
        if tier:
            project_parts.append(f"Tier: {tier}")

        info_style = ParagraphStyle(
            "InfoBox",
            fontSize=8,
            fontName="Helvetica",
            textColor=text_grey,
            leading=11,
        )
        left_cell = Paragraph(
            "<b>BILL TO</b><br/>"
            + "<br/>".join(
                bill_to_parts
                or ["(Customer information)"]
            ),
            info_style,
        )
        right_cell = Paragraph(
            "<b>PROJECT</b><br/>"
            + "<br/>".join(
                project_parts
                or ["(Project details)"]
            ),
            info_style,
        )

        info_table = Table(
            [[left_cell, right_cell]],
            colWidths=[usable_w * 0.5, usable_w * 0.5],
        )
        info_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("BACKGROUND", (0, 0), (-1, -1), brand_light_bg),
            ("BOX", (0, 0), (-1, -1), 0.5, border_grey),
            ("LEFTPADDING", (0, 0), (-1, -1), 10),
            ("RIGHTPADDING", (0, 0), (-1, -1), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 8),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 10))

        # ══════════════════════════════════════
        # OVERVIEW
        # ══════════════════════════════════════

        overview = estimate.get("overview_text") or ""
        if overview:
            elements.append(
                Paragraph(
                    "<b>Overview</b>",
                    section_title_style,
                )
            )
            elements.append(
                Paragraph(
                    overview,
                    ParagraphStyle(
                        "Overview",
                        fontSize=8,
                        fontName="Helvetica",
                        textColor=text_grey,
                        leading=11,
                        spaceBefore=2,
                        spaceAfter=10,
                    ),
                )
            )

        # ══════════════════════════════════════
        # LINE ITEMS: Grouped by section
        # ══════════════════════════════════════

        line_items = estimate.get("line_items", [])
        sections = _group_line_items(line_items)

        col_widths = [
            usable_w * 0.48,   # Description
            usable_w * 0.12,   # Qty
            usable_w * 0.08,   # Unit
            usable_w * 0.15,   # Unit Price
            usable_w * 0.17,   # Total
        ]

        for section in sections:
            # Section header row
            section_header = Table(
                [[
                    Paragraph(
                        section["title"].upper(),
                        ParagraphStyle(
                            "SH",
                            fontSize=8,
                            fontName="Helvetica-Bold",
                            textColor=brand_dark,
                        ),
                    ),
                    "",
                    "",
                    "",
                    "",
                ]],
                colWidths=col_widths,
            )
            section_header.setStyle(TableStyle([
                (
                    "BACKGROUND", (0, 0), (-1, 0),
                    section_bg,
                ),
                (
                    "LINEBELOW", (0, 0), (-1, 0),
                    0.5, brand_accent,
                ),
                ("TOPPADDING", (0, 0), (-1, 0), 5),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 5),
                ("LEFTPADDING", (0, 0), (0, 0), 6),
                ("SPAN", (0, 0), (-1, 0)),
            ]))
            elements.append(section_header)

            # Items
            table_data = []
            for li in section["items"]:
                desc = li.get("description", "")
                notes = li.get("notes", "")
                if notes:
                    desc_p = Paragraph(
                        f"{desc}<br/>"
                        f"<font size=7.5 color='#718096'>"
                        f"{notes}</font>",
                        normal,
                    )
                else:
                    desc_p = Paragraph(desc, normal)

                table_data.append([
                    desc_p,
                    f'{li.get("quantity", 0):.1f}',
                    li.get("unit", ""),
                    f'${li.get("unit_price", 0):,.2f}',
                    f'${li.get("total", 0):,.2f}',
                ])

            # Section subtotal row
            table_data.append([
                Paragraph(
                    f"<b>{section['title']} Subtotal</b>",
                    ParagraphStyle(
                        "SSub",
                        fontSize=8,
                        fontName="Helvetica-Bold",
                        textColor=text_grey,
                        alignment=2,  # RIGHT
                    ),
                ),
                "",
                "",
                "",
                f'${section["total"]:,.2f}',
            ])

            items_table = Table(
                table_data, colWidths=col_widths,
            )

            row_styles = [
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                (
                    "ALIGN", (1, 0), (-1, -1), "RIGHT",
                ),
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
                (
                    "LINEBELOW", (0, -1), (-1, -1),
                    0.5, border_grey,
                ),
                (
                    "TOPPADDING", (0, 0), (-1, -1), 3,
                ),
                (
                    "BOTTOMPADDING", (0, 0), (-1, -1),
                    3,
                ),
                ("LEFTPADDING", (0, 0), (0, -1), 12),
                # Last row (subtotal) styling
                (
                    "FONTNAME", (0, -1), (-1, -1),
                    "Helvetica-Bold",
                ),
                (
                    "LINEABOVE", (0, -1), (-1, -1),
                    1, brand_accent,
                ),
                (
                    "TOPPADDING", (0, -1), (-1, -1),
                    6,
                ),
                (
                    "BOTTOMPADDING", (0, -1), (-1, -1),
                    6,
                ),
                ("SPAN", (0, -1), (3, -1)),
            ]
            # Alternating row backgrounds
            for i in range(0, len(table_data) - 1, 2):
                row_styles.append(
                    (
                        "BACKGROUND", (0, i), (-1, i),
                        colors.HexColor("#fafbfc"),
                    )
                )

            items_table.setStyle(TableStyle(row_styles))
            elements.append(items_table)

        # ══════════════════════════════════════
        # TOTALS
        # ══════════════════════════════════════

        elements.append(Spacer(1, 16))

        subtotal = estimate.get("subtotal", 0)
        overhead = estimate.get("overhead_amount", 0)
        profit = estimate.get("profit_amount", 0)
        total = estimate.get("total", 0)
        oh_pct = estimate.get("overhead_pct", 0.10)
        pr_pct = estimate.get("profit_pct", 0.10)

        totals_data = []
        if oh_pct or pr_pct:
            totals_data.append(
                ["Subtotal", f"${subtotal:,.2f}"]
            )
        if oh_pct:
            totals_data.append([
                f"Overhead ({oh_pct*100:.0f}%)",
                f"${overhead:,.2f}",
            ])
        if pr_pct:
            totals_data.append([
                f"Profit ({pr_pct*100:.0f}%)",
                f"${profit:,.2f}",
            ])
        totals_data.append(
            ["TOTAL", f"${total:,.2f}"]
        )

        tw_label = usable_w * 0.65
        tw_amount = usable_w * 0.35
        totals_table = Table(
            totals_data,
            colWidths=[tw_label, tw_amount],
        )
        totals_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (0, -1), "RIGHT"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("FONTSIZE", (0, 0), (-1, -2), 9),
            (
                "FONTNAME", (0, 0), (0, -2),
                "Helvetica",
            ),
            ("TEXTCOLOR", (0, 0), (-1, -2), text_grey),
            (
                "BOTTOMPADDING", (0, 0), (-1, -1), 4,
            ),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            # Total row
            (
                "FONTNAME", (0, -1), (-1, -1),
                "Helvetica-Bold",
            ),
            ("FONTSIZE", (0, -1), (-1, -1), 14),
            (
                "TEXTCOLOR", (0, -1), (-1, -1),
                brand_dark,
            ),
            (
                "LINEABOVE", (0, -1), (-1, -1),
                1.5, brand_accent,
            ),
            (
                "BACKGROUND", (0, -1), (-1, -1),
                brand_light_bg,
            ),
        ]))
        elements.append(totals_table)

        # ══════════════════════════════════════
        # TERMS & CONDITIONS
        # ══════════════════════════════════════

        elements.append(Spacer(1, 20))
        elements.append(
            HRFlowable(
                width="100%", thickness=0.5,
                color=border_grey,
            )
        )
        elements.append(Spacer(1, 6))
        elements.append(
            Paragraph(
                "TERMS & CONDITIONS",
                section_title_style,
            )
        )

        terms_text = (
            "1. This estimate is valid for 30 days "
            "from the date above.<br/>"
            "2. Payment terms: 50% deposit upon "
            "acceptance, balance due upon "
            "completion.<br/>"
            "3. Any changes to the scope of work "
            "may result in additional charges.<br/>"
            "4. Lead times for cabinet orders are "
            "typically 2-6 weeks depending on tier "
            "and availability.<br/>"
            "5. Prices include standard delivery "
            "and installation unless otherwise "
            "noted.<br/>"
            "6. Warranty: Manufacturer's warranty "
            "applies to all cabinet products. "
            "Installation warranty: 1 year."
        )
        elements.append(Paragraph(terms_text, terms_style))

        # ══════════════════════════════════════
        # ACCEPTANCE / SIGNATURE
        # ══════════════════════════════════════

        if show_signature:
            elements.append(Spacer(1, 24))

            sig_line = "_" * 45
            sig_data = [
                [
                    Paragraph(
                        "<b>ACCEPTED BY:</b>",
                        small_grey,
                    ),
                    "",
                    Paragraph(
                        "<b>AUTHORIZED BY:</b>",
                        small_grey,
                    ),
                    "",
                ],
                [
                    Paragraph(
                        f"Signature: {sig_line}",
                        small_grey,
                    ),
                    "",
                    Paragraph(
                        f"Signature: {sig_line}",
                        small_grey,
                    ),
                    "",
                ],
                [
                    Paragraph(
                        f"Name: {sig_line}",
                        small_grey,
                    ),
                    "",
                    Paragraph(
                        f"Name: {sig_line}",
                        small_grey,
                    ),
                    "",
                ],
                [
                    Paragraph(
                        f"Date: {sig_line}",
                        small_grey,
                    ),
                    "",
                    Paragraph(
                        f"Date: {sig_line}",
                        small_grey,
                    ),
                    "",
                ],
            ]
            sig_table = Table(
                sig_data,
                colWidths=[
                    usable_w * 0.45,
                    usable_w * 0.05,
                    usable_w * 0.45,
                    usable_w * 0.05,
                ],
            )
            sig_table.setStyle(TableStyle([
                ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
                ("TOPPADDING", (0, 0), (-1, -1), 6),
            ]))
            elements.append(sig_table)

        # ══════════════════════════════════════
        # FOOTER: Thank you note
        # ══════════════════════════════════════

        elements.append(Spacer(1, 16))
        elements.append(
            HRFlowable(
                width="100%", thickness=0.5,
                color=border_grey,
            )
        )
        footer_text = (
            "Thank you for the opportunity to "
            "provide this estimate."
        )
        if company.get("name"):
            footer_text += (
                f" — {company['name']}"
            )
        elements.append(Spacer(1, 4))
        elements.append(
            Paragraph(
                f"<i>{footer_text}</i>",
                ParagraphStyle(
                    "Footer",
                    fontSize=8,
                    fontName="Helvetica-Oblique",
                    textColor=text_grey,
                    alignment=1,  # CENTER
                ),
            )
        )

        doc.build(
            elements,
            onLaterPages=_on_later_pages,
        )
        buffer.seek(0)
        return buffer

    # ── Excel Export (for supplier purchase order) ──

    def generate_excel(
        self,
        estimate: Dict[str, Any],
        purchase_order: Dict[str, Any],
    ) -> io.BytesIO:
        """Generate an Excel purchase order."""
        from openpyxl import Workbook
        from openpyxl.styles import (
            Alignment,
            Border,
            Font,
            PatternFill,
            Side,
        )

        wb = Workbook()

        # ── Sheet 1: Purchase Order ──
        ws = wb.active
        ws.title = "Purchase Order"

        header_font = Font(
            bold=True, size=11, color="FFFFFF",
        )
        header_fill = PatternFill(
            start_color="2C3E50",
            end_color="2C3E50",
            fill_type="solid",
        )
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        # Title
        ws["A1"] = "Cabinet Purchase Order"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = (
            f"Property: "
            f"{estimate.get('property_address', '-')}"
        )
        ws["A3"] = (
            f"Date: "
            f"{datetime.now().strftime('%B %d, %Y')}"
        )
        tier = estimate.get("tier", "-")
        ws["A4"] = f"Tier: {tier}"

        # Headers
        headers = [
            "Code", "Type", "Description",
            "Width", "Height", "Qty",
        ]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(
                row=6, column=col, value=header,
            )
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(
                horizontal="center",
            )
            cell.border = thin_border

        # Data
        items = purchase_order.get("items", [])
        for i, item in enumerate(items, 7):
            ws.cell(
                row=i, column=1,
                value=item["code"],
            ).border = thin_border
            ws.cell(
                row=i, column=2,
                value=item["cab_type"],
            ).border = thin_border
            ws.cell(
                row=i, column=3,
                value=item["description"],
            ).border = thin_border
            ws.cell(
                row=i, column=4,
                value=f'{item["width_inches"]}"',
            ).border = thin_border
            ws.cell(
                row=i, column=5,
                value=f'{item["height_inches"]}"',
            ).border = thin_border
            qty_cell = ws.cell(
                row=i, column=6, value=item["qty"],
            )
            qty_cell.border = thin_border
            qty_cell.alignment = Alignment(
                horizontal="center",
            )

        # Total row
        total_row = 7 + len(items)
        ws.cell(
            row=total_row, column=5,
            value="Total Boxes:",
        ).font = Font(bold=True)
        ws.cell(
            row=total_row, column=6,
            value=purchase_order.get("total_boxes", 0),
        ).font = Font(bold=True)

        # Column widths
        ws.column_dimensions["A"].width = 12
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 40
        ws.column_dimensions["D"].width = 10
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 8

        # ── Sheet 2: Estimate Summary ──
        ws2 = wb.create_sheet("Estimate Summary")

        ws2["A1"] = "Cabinet Estimate Summary"
        ws2["A1"].font = Font(bold=True, size=14)

        line_items = estimate.get("line_items", [])
        li_headers = [
            "Description", "Qty", "Unit",
            "Unit Price", "Total",
        ]
        for col, header in enumerate(li_headers, 1):
            cell = ws2.cell(
                row=3, column=col, value=header,
            )
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border

        for i, li in enumerate(line_items, 4):
            ws2.cell(
                row=i, column=1,
                value=li.get("description", ""),
            ).border = thin_border
            ws2.cell(
                row=i, column=2,
                value=round(li.get("quantity", 0), 2),
            ).border = thin_border
            ws2.cell(
                row=i, column=3,
                value=li.get("unit", ""),
            ).border = thin_border
            ws2.cell(
                row=i, column=4,
                value=round(
                    li.get("unit_price", 0), 2,
                ),
            ).border = thin_border
            ws2.cell(
                row=i, column=5,
                value=round(li.get("total", 0), 2),
            ).border = thin_border

        sr = 4 + len(line_items) + 1
        ws2.cell(
            row=sr, column=4, value="Subtotal:",
        ).font = Font(bold=True)
        ws2.cell(
            row=sr, column=5,
            value=estimate.get("subtotal", 0),
        )
        ws2.cell(
            row=sr + 1, column=4, value="Overhead:",
        ).font = Font(bold=True)
        ws2.cell(
            row=sr + 1, column=5,
            value=estimate.get("overhead_amount", 0),
        )
        ws2.cell(
            row=sr + 2, column=4, value="Profit:",
        ).font = Font(bold=True)
        ws2.cell(
            row=sr + 2, column=5,
            value=estimate.get("profit_amount", 0),
        )
        ws2.cell(
            row=sr + 3, column=4, value="TOTAL:",
        ).font = Font(bold=True, size=12)
        ws2.cell(
            row=sr + 3, column=5,
            value=estimate.get("total", 0),
        ).font = Font(bold=True, size=12)

        ws2.column_dimensions["A"].width = 40
        ws2.column_dimensions["B"].width = 10
        ws2.column_dimensions["C"].width = 8
        ws2.column_dimensions["D"].width = 12
        ws2.column_dimensions["E"].width = 12

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
